import csv
import io
import pathlib
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path('/root/pa-clients/jivo-data-bank')
LOC_DIR = ROOT / 'ecom/locations'
OUT_DIR = ROOT / 'reports/amazon-now-exact-16-states-availability'
OUT_XLSX = OUT_DIR / 'amazon-now-exact-16-states-availability.xlsx'
PLATFORM = 'amazon-now'
STATES = [
    'Delhi', 'Maharashtra', 'Karnataka', 'Haryana', 'Uttar Pradesh', 'Telangana',
    'West Bengal', 'Punjab', 'Tamil Nadu', 'Andhra Pradesh', 'Kerala', 'Odisha',
    'Gujarat', 'Rajasthan', 'Madhya Pradesh', 'Chandigarh'
]
CITY_STATE = {
    'Delhi': 'Delhi',
    'Mumbai': 'Maharashtra', 'Pune': 'Maharashtra', 'Nashik': 'Maharashtra', 'Nagpur': 'Maharashtra',
    'Bengaluru': 'Karnataka', 'Mysuru': 'Karnataka',
    'Gurgaon': 'Haryana', 'Gurugram': 'Haryana', 'Faridabad': 'Haryana',
    'Noida': 'Uttar Pradesh', 'Ghaziabad': 'Uttar Pradesh', 'Lucknow': 'Uttar Pradesh',
    'Hyderabad': 'Telangana',
    'Kolkata': 'West Bengal',
    'Ludhiana': 'Punjab',
    'Chennai': 'Tamil Nadu', 'Coimbatore': 'Tamil Nadu',
    'Vijayawada': 'Andhra Pradesh',
    'Kochi': 'Kerala',
    'Ahmedabad': 'Gujarat', 'Surat': 'Gujarat', 'Vadodara': 'Gujarat',
    'Jaipur': 'Rajasthan',
    'Bhopal': 'Madhya Pradesh', 'Indore': 'Madhya Pradesh',
    'Chandigarh': 'Chandigarh',
}
OUT_DIR.mkdir(parents=True, exist_ok=True)

coverage_by_state = {st: set() for st in STATES}
coverage_by_city = defaultdict(set)
city_state = {}
skus = set()
best = {}
source_cities = set()
latest_date = ''

for path in sorted(LOC_DIR.glob('*.md')):
    city = path.stem
    state = CITY_STATE.get(city)
    if state not in STATES:
        continue
    text = path.read_text(errors='ignore')
    if f',{PLATFORM},' not in text or '```csv' not in text:
        continue
    csv_block = text.split('```csv', 1)[1].split('```', 1)[0].strip()
    city_had_rows = False
    for r in csv.DictReader(io.StringIO(csv_block)):
        if r.get('platform') != PLATFORM:
            continue
        p = str(r.get('pincode') or '').strip()
        sku = (r.get('canonical_sku') or '').strip()
        if not p or not sku:
            continue
        city_had_rows = True
        source_cities.add(city)
        skus.add(sku)
        coverage_by_state[state].add(p)
        coverage_by_city[city].add(p)
        city_state[city] = state
        stock = 1 if r.get('in_stock') in (1, True, '1', 'true', 'True') else 0
        date = str(r.get('date_ist') or '')
        run = str(r.get('run_id') or '')
        latest_date = max(latest_date, date)
        key = (p, sku)
        rank = (date, run, stock)
        prev = best.get(key)
        if prev is None or rank > prev['rank']:
            best[key] = {'stock': stock, 'rank': rank, 'city': city, 'state': state}

skus = sorted(skus)
cities = sorted(coverage_by_city.keys(), key=lambda c: (STATES.index(city_state[c]), c))


def title_from_slug(slug: str) -> str:
    out = slug.replace('jivo-', 'JIVO ', 1).replace('-', ' ').title()
    for a, b in [('1L','1L'),('2L','2L'),('3L','3L'),('4L','4L'),('5L','5L'),('15L','15L'),('200Ml','200ml'),('500Ml','500ml')]:
        out = out.replace(a, b)
    return out


def pct(avail, total):
    return avail / total if total else 0

state_counts = {sku: {} for sku in skus}
city_counts = {sku: {} for sku in skus}
for sku in skus:
    for st in STATES:
        pins = coverage_by_state[st]
        avail = sum(1 for p in pins if best.get((p, sku), {}).get('stock') == 1)
        state_counts[sku][st] = (avail, len(pins))
    for city in cities:
        pins = coverage_by_city[city]
        avail = sum(1 for p in pins if best.get((p, sku), {}).get('stock') == 1)
        city_counts[sku][city] = (avail, len(pins))

wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill('solid', fgColor='111827')
header_font = Font(color='FFFFFF', bold=True)
good_fill = PatternFill('solid', fgColor='DCFCE7')
mid_fill = PatternFill('solid', fgColor='FEF3C7')
bad_fill = PatternFill('solid', fgColor='FEE2E2')
zero_fill = PatternFill('solid', fgColor='F3F4F6')
thin = Side(style='thin', color='D1D5DB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill_for(avail, total):
    v = pct(avail, total)
    if total == 0 or avail == 0:
        return zero_fill
    if v >= 0.8:
        return good_fill
    if v >= 0.4:
        return mid_fill
    return bad_fill


def style_matrix(ws, columns, is_pct=True):
    ws.freeze_panes = 'B2'
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 36
    ws.column_dimensions['A'].width = 66
    for idx in range(2, len(columns) + 2):
        ws.column_dimensions[get_column_letter(idx)].width = 16
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[0].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if is_pct:
            for cell in row[1:]:
                cell.number_format = '0.0%'
                v = cell.value or 0
                cell.fill = zero_fill if v == 0 else good_fill if v >= 0.8 else mid_fill if v >= 0.4 else bad_fill


def add_percent_sheet(name, columns, counts):
    ws = wb.create_sheet(name)
    ws.append(['SKU'] + columns)
    for sku in skus:
        ws.append([title_from_slug(sku)] + [pct(*counts[sku][col]) for col in columns])
    style_matrix(ws, columns, True)


def add_count_sheet(name, columns, counts):
    ws = wb.create_sheet(name)
    ws.append(['SKU'] + columns)
    for sku in skus:
        ws.append([title_from_slug(sku)] + [f'{counts[sku][col][0]}/{counts[sku][col][1]}' for col in columns])
    style_matrix(ws, columns, False)
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            a, t = map(int, str(cell.value).split('/'))
            cell.fill = fill_for(a, t)

add_percent_sheet('State Percent Matrix', STATES, state_counts)
add_count_sheet('State Counts', STATES, state_counts)
add_percent_sheet('City Percent Matrix', cities, city_counts)
add_count_sheet('City Counts', cities, city_counts)

ws = wb.create_sheet('Long Format')
ws.append(['Platform', 'State', 'City', 'SKU Slug', 'SKU', 'Available Pincodes', 'Serviceable Pincodes', 'Availability %'])
for sku in skus:
    for st in STATES:
        a, t = state_counts[sku][st]
        ws.append([PLATFORM, st, '', sku, title_from_slug(sku), a, t, pct(a, t)])
for cell in ws[1]:
    cell.fill = header_fill; cell.font = header_font; cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row[7].number_format = '0.0%'
ws.freeze_panes = 'A2'
for col, width in {'A':16,'B':22,'C':18,'D':74,'E':74,'F':18,'G':20,'H':16}.items():
    ws.column_dimensions[col].width = width

ws = wb.create_sheet('Source')
source_rows = [
    ('Platform', PLATFORM),
    ('Data source used', 'ecom/locations/*.md city-hub CSV observations from approved jivo-data-bank repo'),
    ('Reason not website data.js SKU rows', 'reports/ecom-availability-app/data.js has Amazon Now coverage rows but 0 Amazon Now SKU availability rows'),
    ('Latest observation date', latest_date or 'not found'),
    ('Requested states', ', '.join(STATES)),
    ('State count', len(STATES)),
    ('SKU count', len(skus)),
    ('Source city hubs', ', '.join(sorted(source_cities))),
    ('States with 0/0', ', '.join([st for st in STATES if len(coverage_by_state[st]) == 0]) or 'None'),
    ('Calculation', 'Available pincodes / Amazon Now serviceable pincodes from latest city-hub observations'),
    ('Availability numerator', 'Latest Amazon Now observation per (pincode, SKU) with in_stock=1'),
]
for row in source_rows:
    ws.append(row)
for cell in ws[1]:
    cell.fill = header_fill; cell.font = header_font
for row in ws.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.column_dimensions['A'].width = 34
ws.column_dimensions['B'].width = 130

wb.save(OUT_XLSX)
print(OUT_XLSX)
print('sku_count', len(skus))
print('states', STATES)
print('zero_states', [st for st in STATES if len(coverage_by_state[st]) == 0])
print('source_cities', sorted(source_cities))
