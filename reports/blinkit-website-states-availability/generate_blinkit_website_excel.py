import json
import pathlib
import re
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path('/root/pa-clients/jivo-data-bank')
DATA_JS = ROOT / 'reports/ecom-availability-app/data.js'
OUT_DIR = ROOT / 'reports/blinkit-website-states-availability'
OUT_XLSX = OUT_DIR / 'blinkit-website-states-availability.xlsx'
PLATFORM = 'blinkit'

text = DATA_JS.read_text()
raw = text.split('=', 1)[1].strip().rstrip(';')
data = json.loads(raw)

# Use only states present in the website data / coverage for Blinkit.
coverage_rows = [c for c in data['coverage'] if c.get('platform') == PLATFORM]
states = [c['state'] for c in coverage_rows]
# User-requested additions beyond the website state list.
# Punjab is filled from the repo data bank city hub; Mizoram is included if requested,
# but no Blinkit pincode availability source exists for it in the approved data bank.
for extra_state in ['Punjab', 'Mizoram']:
    if extra_state not in states:
        states.append(extra_state)
coverage_by_state = {c['state']: set(map(str, c.get('pincodes', []))) for c in coverage_rows}
coverage_by_state.setdefault('Punjab', set())
coverage_by_state.setdefault('Mizoram', set())

# City map from pincode metadata and records.
pincode_meta = data.get('pincodes', {})
pincode_city = {}
pincode_state = {}
for p, meta in pincode_meta.items():
    if isinstance(meta, dict):
        pincode_city[str(p)] = meta.get('city') or meta.get('c') or ''
        pincode_state[str(p)] = meta.get('state') or meta.get('st') or ''
for r in data['records']:
    if r.get('pl') == PLATFORM:
        p = str(r.get('p'))
        pincode_city[p] = r.get('c') or pincode_city.get(p, '')
        pincode_state[p] = r.get('st') or pincode_state.get(p, '')

# SKU rows: all Blinkit SKUs appearing on website data, excluding coverage marker.
skus = sorted({r.get('s') for r in data['records'] if r.get('pl') == PLATFORM and r.get('s') and r.get('s') != '__coverage__'})

def title_from_slug(slug):
    return slug.replace('jivo-', 'JIVO ', 1).replace('-', ' ').title().replace('1L', '1L').replace('5L', '5L').replace('2L', '2L')

# Latest/best availability per (pincode, sku). Prefer in-stock; otherwise latest date/run.
best = {}
for r in data['records']:
    if r.get('pl') != PLATFORM or r.get('s') == '__coverage__':
        continue
    p = str(r.get('p'))
    s = r.get('s')
    key = (p, s)
    stock = 1 if r.get('stock') in (1, True, '1', 'true', 'True') else 0
    date = str(r.get('date') or '')
    run = str(r.get('run') or '')
    rank = (stock, date, run)
    prev = best.get(key)
    if prev is None or rank > prev['rank']:
        best[key] = {'stock': stock, 'rank': rank, 'city': r.get('c') or pincode_city.get(p, ''), 'state': r.get('st') or pincode_state.get(p, '')}

# Add Punjab availability from the data bank city hub (not present in the deployed website data layer).
# Source: ecom/locations/Ludhiana.md, latest Blinkit observation per (pincode, SKU).
punjab_source = ROOT / 'ecom/locations/Ludhiana.md'
punjab_city = 'Ludhiana'
punjab_latest_date = ''
if punjab_source.exists():
    import csv
    import io
    ludhiana_text = punjab_source.read_text()
    if '```csv' in ludhiana_text:
        pin_section = ludhiana_text.split('## Pincodes', 1)[1].split('## SKUs', 1)[0]
        ludhiana_pins = set(re.findall(r'\[\[(\d{6})\]\]', pin_section))
        coverage_by_state['Punjab'] = ludhiana_pins
        for p in ludhiana_pins:
            pincode_city[p] = punjab_city
            pincode_state[p] = 'Punjab'
        csv_block = ludhiana_text.split('```csv', 1)[1].split('```', 1)[0].strip()
        for r in csv.DictReader(io.StringIO(csv_block)):
            if r.get('platform') != PLATFORM or r.get('canonical_sku') not in skus:
                continue
            p = str(r.get('pincode'))
            s = r.get('canonical_sku')
            stock = 1 if r.get('in_stock') in (1, True, '1', 'true', 'True') else 0
            date = str(r.get('date_ist') or '')
            run = str(r.get('run_id') or '')
            punjab_latest_date = max(punjab_latest_date, date)
            key = (p, s)
            # Use date/run as the latest selector for the data-bank city hub.
            rank = (date, run, stock)
            prev = best.get(key)
            if prev is None or rank > prev['rank']:
                best[key] = {'stock': stock, 'rank': rank, 'city': punjab_city, 'state': 'Punjab'}

# City coverage: platform-serviceable pincodes grouped by city from the Blinkit coverage states.
coverage_by_city = defaultdict(set)
city_state = {}
for st, pins in coverage_by_state.items():
    for p in pins:
        city = pincode_city.get(p) or 'Unknown'
        coverage_by_city[city].add(p)
        city_state[city] = st
cities = sorted(coverage_by_city.keys())

# Calculations.
state_counts = {}
city_counts = {}
for sku in skus:
    state_counts[sku] = {}
    for st in states:
        pins = coverage_by_state[st]
        avail = sum(1 for p in pins if best.get((p, sku), {}).get('stock') == 1)
        state_counts[sku][st] = (avail, len(pins))
    city_counts[sku] = {}
    for city in cities:
        pins = coverage_by_city[city]
        avail = sum(1 for p in pins if best.get((p, sku), {}).get('stock') == 1)
        city_counts[sku][city] = (avail, len(pins))

wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill('solid', fgColor='111827')
header_font = Font(color='FFFFFF', bold=True)
sub_fill = PatternFill('solid', fgColor='E5E7EB')
good_fill = PatternFill('solid', fgColor='DCFCE7')
mid_fill = PatternFill('solid', fgColor='FEF3C7')
bad_fill = PatternFill('solid', fgColor='FEE2E2')
zero_fill = PatternFill('solid', fgColor='F3F4F6')
thin = Side(style='thin', color='D1D5DB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def pct(avail, total):
    return (avail / total) if total else 0


def style_matrix(ws, first_col_label, columns, is_pct=True):
    ws.freeze_panes = 'B2'
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row[0].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        row[0].font = Font(bold=True)
    ws.column_dimensions['A'].width = 46
    for idx in range(2, len(columns) + 2):
        ws.column_dimensions[get_column_letter(idx)].width = 16
    ws.row_dimensions[1].height = 36
    if is_pct:
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                v = cell.value or 0
                cell.number_format = '0.0%'
                if v == 0:
                    cell.fill = zero_fill
                elif v >= 0.8:
                    cell.fill = good_fill
                elif v >= 0.4:
                    cell.fill = mid_fill
                else:
                    cell.fill = bad_fill


def add_percent_sheet(name, columns, counts):
    ws = wb.create_sheet(name)
    ws.append(['SKU'] + columns)
    for sku in skus:
        ws.append([title_from_slug(sku)] + [pct(*counts[sku][col]) for col in columns])
    style_matrix(ws, 'SKU', columns, True)
    return ws


def add_count_sheet(name, columns, counts):
    ws = wb.create_sheet(name)
    ws.append(['SKU'] + columns)
    for sku in skus:
        ws.append([title_from_slug(sku)] + [f'{counts[sku][col][0]}/{counts[sku][col][1]}' for col in columns])
    style_matrix(ws, 'SKU', columns, False)
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            a, t = map(int, str(cell.value).split('/'))
            cell.fill = zero_fill if t == 0 or a == 0 else (good_fill if a / t >= 0.8 else mid_fill if a / t >= 0.4 else bad_fill)
    return ws

add_percent_sheet('State Percent Matrix', states, state_counts)
add_count_sheet('State Counts', states, state_counts)
add_percent_sheet('City Percent Matrix', cities, city_counts)
add_count_sheet('City Counts', cities, city_counts)

ws = wb.create_sheet('Long Format')
ws.append(['Platform', 'State', 'City', 'SKU Slug', 'SKU', 'Available Pincodes', 'Serviceable Pincodes', 'Availability %'])
for sku in skus:
    for st in states:
        a, t = state_counts[sku][st]
        ws.append([PLATFORM, st, '', sku, title_from_slug(sku), a, t, pct(a, t)])
for cell in ws[1]:
    cell.fill = header_fill; cell.font = header_font; cell.border = border
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
    row[7].number_format = '0.0%'
ws.freeze_panes = 'A2'
for col, width in {'A':14,'B':22,'C':18,'D':48,'E':48,'F':18,'G':20,'H':16}.items():
    ws.column_dimensions[col].width = width

ws = wb.create_sheet('Source')
source_rows = [
    ('Website requested', 'https://ecom-availability-app.vercel.app'),
    ('Data file used', 'reports/ecom-availability-app/data.js'),
    ('Generated from', data.get('generatedFrom')),
    ('Generated at', data.get('generatedAt')),
    ('Latest observation date', data.get('latestObservationDate')),
    ('Platform filter', PLATFORM),
    ('States included', ', '.join(states)),
    ('State count', len(states)),
    ('SKU count', len(skus)),
    ('Calculation', 'Available pincodes / Blinkit serviceable pincodes'),
    ('Serviceable denominator', 'Website coverage[] pincodes for website states; Punjab uses ecom/locations/Ludhiana.md Blinkit pincodes from the data bank'),
    ('Availability numerator', 'Website records[] rows for website states; Punjab uses latest Blinkit observation per (pincode, SKU) from ecom/locations/Ludhiana.md with in_stock=1'),
    ('Punjab source', f'ecom/locations/Ludhiana.md; latest Blinkit observation date {punjab_latest_date or "not found"}; 47 serviceable pincodes'),
    ('Mizoram source check', 'No Blinkit pincode availability source found for Mizoram in reports/ecom-availability-app/data.js or ecom/locations; Mizoram is included as 0/0.'),
    ('Note', 'Punjab was added from the data bank because it is not present in the deployed website data layer for Blinkit. Mizoram is included on request, but no Blinkit availability data was found in the approved data bank.'),
]
for row in source_rows:
    ws.append(row)
for cell in ws[1]:
    cell.fill = header_fill; cell.font = header_font
for row in ws.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 110

OUT_DIR.mkdir(parents=True, exist_ok=True)
wb.save(OUT_XLSX)

# Verify workbook can be loaded.
wb2 = load_workbook(OUT_XLSX, read_only=True, data_only=True)
assert wb2.sheetnames == ['State Percent Matrix', 'State Counts', 'City Percent Matrix', 'City Counts', 'Long Format', 'Source']
assert wb2['State Percent Matrix'].max_column - 1 == len(states)
assert wb2['State Percent Matrix'].max_row - 1 == len(skus)
wb2.close()

print('created', OUT_XLSX)
print('states', len(states), states)
print('skus', len(skus), skus)
print('size', OUT_XLSX.stat().st_size)
