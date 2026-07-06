from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = Path('/root/.hermes/document_cache/doc_aa130968b235_blinkit-website-states-availability-3.xlsx')
OUT_DIR = Path('/root/pa-clients/jivo-data-bank/reports/instamart-exact-16-states-availability')
OUT = OUT_DIR / 'instamart-exact-16-states-availability.xlsx'
OUT_DIR.mkdir(parents=True, exist_ok=True)

STATES = [
    'Delhi', 'Maharashtra', 'Karnataka', 'Haryana', 'Uttar Pradesh', 'Telangana',
    'West Bengal', 'Punjab', 'Tamil Nadu', 'Andhra Pradesh', 'Kerala', 'Odisha',
    'Gujarat', 'Rajasthan', 'Madhya Pradesh', 'Chandigarh'
]

wb_src = load_workbook(SRC, data_only=False)
skus = [wb_src['State Percent Matrix'].cell(r, 1).value for r in range(2, wb_src['State Percent Matrix'].max_row + 1)]
skus = [s for s in skus if s]
wb_src.close()

header_fill = PatternFill('solid', fgColor='111827')
header_font = Font(color='FFFFFF', bold=True)
zero_fill = PatternFill('solid', fgColor='F3F4F6')
thin = Side(style='thin', color='D1D5DB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = load_workbook(SRC)
# Remove all sheets and rebuild with same sheet names/order for clean Instamart output.
for s in list(wb.sheetnames):
    del wb[s]


def style_matrix(ws, columns, is_pct):
    ws.freeze_panes = 'B2'
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 36
    ws.column_dimensions['A'].width = 46
    for idx in range(2, len(columns) + 2):
        ws.column_dimensions[get_column_letter(idx)].width = 16
    for row in ws.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[0].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if cell.column > 1:
                cell.fill = zero_fill
                if is_pct:
                    cell.number_format = '0.0%'


ws = wb.create_sheet('State Percent Matrix')
ws.append(['SKU'] + STATES)
for sku in skus:
    ws.append([sku] + [0 for _ in STATES])
style_matrix(ws, STATES, True)

ws = wb.create_sheet('State Counts')
ws.append(['SKU'] + STATES)
for sku in skus:
    ws.append([sku] + ['0/0' for _ in STATES])
style_matrix(ws, STATES, False)

# No Instamart city coverage exists in approved source, so keep city matrix as SKU-only.
ws = wb.create_sheet('City Percent Matrix')
ws.append(['SKU'])
for sku in skus:
    ws.append([sku])
style_matrix(ws, [], True)

ws = wb.create_sheet('City Counts')
ws.append(['SKU'])
for sku in skus:
    ws.append([sku])
style_matrix(ws, [], False)

ws = wb.create_sheet('Long Format')
ws.append(['Platform', 'State', 'City', 'SKU Slug', 'SKU', 'Available Pincodes', 'Serviceable Pincodes', 'Availability %'])
for sku in skus:
    slug = str(sku).lower().replace(' ', '-').replace('jivo-', 'jivo-')
    for st in STATES:
        ws.append(['instamart', st, '', slug, sku, 0, 0, 0])
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row[7].number_format = '0.0%'
ws.freeze_panes = 'A2'
for col, width in {'A':14,'B':22,'C':18,'D':48,'E':48,'F':18,'G':20,'H':16}.items():
    ws.column_dimensions[col].width = width

ws = wb.create_sheet('Source')
source_rows = [
    ('Platform', 'instamart'),
    ('Requested format', 'Same as shared Blinkit workbook'),
    ('Requested states', ', '.join(STATES)),
    ('State count', len(STATES)),
    ('SKU rows copied from shared workbook', len(skus)),
    ('Approved source checked', 'reports/ecom-availability-app/data.js and data-freshness.json'),
    ('Result', 'No Instamart availability coverage/data files found in the approved Jivo GitHub repo'),
    ('Freshness status', 'instamart is RED / known_dead=true / no data files / no baseline samples / no history.csv'),
    ('Calculation', 'All values shown as 0/0 and 0.0% because no Instamart serviceable pincode denominator exists in the approved source'),
    ('Important note', 'These are not real zero availability observations; they mean Instamart data is absent from the approved repo source.'),
]
for row in source_rows:
    ws.append(row)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
for row in ws.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.column_dimensions['A'].width = 34
ws.column_dimensions['B'].width = 120

wb.save(OUT)
print(OUT)
print('sku_rows', len(skus))
print('states', STATES)
