import csv
import io
import json
import pathlib
import re
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path('/root/pa-clients/jivo-data-bank')
DATA_JS = ROOT / 'reports/ecom-availability-app/data.js'
RANKING_XLSX = ROOT / 'reports/june-pomace-litres-ranking/june-pomace-litres-ranking.xlsx'
OUT_DIR = ROOT / 'reports/june-pomace-blinkit-availability'
OUT_XLSX = OUT_DIR / 'june-pomace-state-ranking-blinkit-availability.xlsx'
PLATFORM = 'blinkit'
POMACE_1L = 'jivo-pomace-olive-oil-1l'

OUT_DIR.mkdir(parents=True, exist_ok=True)

def load_data_js():
    text = DATA_JS.read_text()
    raw = text.split('=', 1)[1].strip().rstrip(';')
    return json.loads(raw)

def stock_value(v):
    return 1 if v in (1, True, '1', 'true', 'True') else 0

def pct(a, t):
    return a / t if t else 0

def title_from_slug(slug):
    return slug.replace('jivo-', 'JIVO ', 1).replace('-', ' ').title().replace('1L', '1L').replace('2L', '2L').replace('5L', '5L')

# State ranking source from the already verified live-app report.
wb_rank = load_workbook(RANKING_XLSX, read_only=True, data_only=True)
ws_rank = wb_rank['State Ranking']
ranking_rows = []
for row in ws_rank.iter_rows(min_row=2, values_only=True):
    if not row or row[0] is None:
        continue
    ranking_rows.append({'rank': int(row[0]), 'state': str(row[1]), 'litres': float(row[2] or 0)})
wb_rank.close()
ranking_state_set = {r['state'] for r in ranking_rows}

# Website data layer.
data = load_data_js()
coverage_rows = [c for c in data['coverage'] if c.get('platform') == PLATFORM]
coverage_by_state = {c['state']: set(map(str, c.get('pincodes', []))) for c in coverage_rows}
pincode_meta = data.get('pincodes', {})
pincode_city = {}
pincode_state = {}
for p, meta in pincode_meta.items():
    if isinstance(meta, dict):
        pincode_city[str(p)] = meta.get('city') or meta.get('c') or ''
        pincode_state[str(p)] = meta.get('state') or meta.get('st') or ''
for r in data['records']:
    if r.get('pl') == PLATFORM:
        p = str(r.get('p'))
        pincode_city[p] = r.get('c') or pincode_city.get(p, '')
        pincode_state[p] = r.get('st') or pincode_state.get(p, '')

skus = sorted({r.get('s') for r in data['records'] if r.get('pl') == PLATFORM and r.get('s') and r.get('s') != '__coverage__'})

best = {}
for r in data['records']:
    if r.get('pl') != PLATFORM or r.get('s') == '__coverage__':
        continue
    p = str(r.get('p'))
    s = r.get('s')
    stock = stock_value(r.get('stock'))
    date = str(r.get('date') or '')
    run = str(r.get('run') or '')
    # For website data, prefer in-stock if there are multiple observations, then latest.
    rank = (stock, date, run)
    key = (p, s)
    prev = best.get(key)
    if prev is None or rank > prev['rank']:
        best[key] = {'stock': stock, 'rank': rank, 'city': r.get('c') or pincode_city.get(p, ''), 'state': r.get('st') or pincode_state.get(p, ''), 'source': 'website data.js'}

# Backfill Punjab from approved repo city hub, since Blinkit Ludhiana data is not present in deployed website coverage.
punjab_source = ROOT / 'ecom/locations/Ludhiana.md'
punjab_latest_date = ''
if punjab_source.exists():
    text = punjab_source.read_text()
    try:
        pin_section = text.split('## Pincodes', 1)[1].split('## SKUs', 1)[0]
        pins = set(re.findall(r'\[\[(\d{6})\]\]', pin_section))
        if pins:
            coverage_by_state['Punjab'] = pins
            for p in pins:
                pincode_city[p] = 'Ludhiana'
                pincode_state[p] = 'Punjab'
    except Exception:
        pins = set()
    if '```csv' in text:
        csv_block = text.split('```csv', 1)[1].split('```', 1)[0].strip()
        for r in csv.DictReader(io.StringIO(csv_block)):
            if r.get('platform') != PLATFORM:
                continue
            s = r.get('canonical_sku')
            if not s:
                continue
            if s not in skus:
                skus.append(s)
            p = str(r.get('pincode'))
            stock = stock_value(r.get('in_stock'))
            date = str(r.get('date_ist') or '')
            run = str(r.get('run_id') or '')
            punjab_latest_date = max(punjab_latest_date, date)
            key = (p, s)
            # For the data-bank city hub, choose latest observation first.
            rank = (date, run, stock)
            prev = best.get(key)
            if prev is None or rank > prev['rank']:
                best[key] = {'stock': stock, 'rank': rank, 'city': 'Ludhiana', 'state': 'Punjab', 'source': 'ecom/locations/Ludhiana.md'}
skus = sorted(skus)

# Keep only states from the June Pomace ranking where Blinkit serves at least one pincode.
served_states = {st for st, pins in coverage_by_state.items() if pins}
filtered_ranking = [r for r in ranking_rows if r['state'] in served_states]
removed_ranking = [r for r in ranking_rows if r['state'] not in served_states]
states = [r['state'] for r in filtered_ranking]

# Counts.
state_sku_counts = {sku: {} for sku in skus}
for sku in skus:
    for st in states:
        pins = coverage_by_state.get(st, set())
        a = sum(1 for p in pins if best.get((p, sku), {}).get('stock') == 1)
        state_sku_counts[sku][st] = (a, len(pins))

def any_jivo_counts(st):
    pins = coverage_by_state.get(st, set())
    available = 0
    for p in pins:
        if any(best.get((p, sku), {}).get('stock') == 1 for sku in skus):
            available += 1
    return available, len(pins)

# Styling.
header_fill = PatternFill('solid', fgColor='111827')
header_font = Font(color='FFFFFF', bold=True)
sub_fill = PatternFill('solid', fgColor='E5E7EB')
good_fill = PatternFill('solid', fgColor='DCFCE7')
mid_fill = PatternFill('solid', fgColor='FEF3C7')
bad_fill = PatternFill('solid', fgColor='FEE2E2')
zero_fill = PatternFill('solid', fgColor='F3F4F6')
thin = Side(style='thin', color='D1D5DB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def color_for_pct(v, total):
    if total == 0 or v == 0:
        return zero_fill
    if v >= 0.8:
        return good_fill
    if v >= 0.4:
        return mid_fill
    return bad_fill

def style_header(ws):
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 38

wb = Workbook()
ws = wb.active
ws.title = 'Filtered State Ranking'
ws.append([
    'June Pomace Rank', 'State', 'June Pomace Litres Sold', 'Blinkit Serviceable Pincodes',
    'JIVO Pomace Olive Oil 1L Available Pincodes', 'JIVO Pomace Olive Oil 1L Availability %',
    'Any JIVO SKU Available Pincodes', 'Any JIVO SKU Availability %'
])
for r in filtered_ranking:
    st = r['state']
    pom_a, pom_t = state_sku_counts.get(POMACE_1L, {}).get(st, (0, len(coverage_by_state.get(st, set()))))
    any_a, any_t = any_jivo_counts(st)
    ws.append([r['rank'], st, r['litres'], pom_t, pom_a, pct(pom_a, pom_t), any_a, pct(any_a, any_t)])
style_header(ws)
ws.freeze_panes = 'A2'
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row[1].alignment = Alignment(horizontal='left')
    row[5].number_format = '0.0%'
    row[7].number_format = '0.0%'
    row[5].fill = color_for_pct(row[5].value or 0, row[3].value or 0)
    row[7].fill = color_for_pct(row[7].value or 0, row[3].value or 0)
for col, width in {'A':16,'B':26,'C':20,'D':22,'E':28,'F':24,'G':26,'H':24}.items():
    ws.column_dimensions[col].width = width

# SKU percent matrix, for all Blinkit JIVO SKUs in the served states.
ws = wb.create_sheet('JIVO SKU Percent Matrix')
ws.append(['SKU'] + states)
for sku in skus:
    ws.append([title_from_slug(sku)] + [pct(*state_sku_counts[sku][st]) for st in states])
style_header(ws)
ws.freeze_panes = 'B2'
ws.column_dimensions['A'].width = 48
for idx in range(2, len(states) + 2):
    ws.column_dimensions[get_column_letter(idx)].width = 16
for row in ws.iter_rows(min_row=2):
    row[0].font = Font(bold=True)
    row[0].alignment = Alignment(horizontal='left', wrap_text=True)
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for idx, cell in enumerate(row[1:], start=0):
        st = states[idx]
        cell.number_format = '0.0%'
        cell.fill = color_for_pct(cell.value or 0, len(coverage_by_state.get(st, set())))

ws = wb.create_sheet('JIVO SKU Counts')
ws.append(['SKU'] + states)
for sku in skus:
    ws.append([title_from_slug(sku)] + [f'{state_sku_counts[sku][st][0]}/{state_sku_counts[sku][st][1]}' for st in states])
style_header(ws)
ws.freeze_panes = 'B2'
ws.column_dimensions['A'].width = 48
for idx in range(2, len(states) + 2):
    ws.column_dimensions[get_column_letter(idx)].width = 16
for row in ws.iter_rows(min_row=2):
    row[0].font = Font(bold=True)
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in row[1:]:
        a, t = map(int, str(cell.value).split('/'))
        cell.fill = color_for_pct(a/t if t else 0, t)

ws = wb.create_sheet('Removed States')
ws.append(['June Pomace Rank', 'State', 'June Pomace Litres Sold', 'Reason Removed'])
for r in removed_ranking:
    ws.append([r['rank'], r['state'], r['litres'], 'Blinkit has 0 serviceable pincodes in the approved availability data used'])
style_header(ws)
ws.freeze_panes = 'A2'
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row[1].alignment = Alignment(horizontal='left')
for col, width in {'A':16,'B':34,'C':20,'D':62}.items():
    ws.column_dimensions[col].width = width

ws = wb.create_sheet('Source')
source_rows = [
    ('Sales ranking source', 'reports/june-pomace-litres-ranking/june-pomace-litres-ranking.xlsx'),
    ('Sales ranking API source', '/api/dashboard/state-sales; year=2026; month=6; metric=litres; category=OLIVE; sub_category=JIVO POMACE; brand=JIVO'),
    ('Availability platform', PLATFORM),
    ('Availability website data file', 'reports/ecom-availability-app/data.js'),
    ('Website data generatedFrom', data.get('generatedFrom')),
    ('Website data generatedAt', data.get('generatedAt')),
    ('Website latest observation date', data.get('latestObservationDate')),
    ('Punjab backfill source', 'ecom/locations/Ludhiana.md'),
    ('Punjab latest Blinkit observation date', punjab_latest_date or 'not found'),
    ('Included states count', len(states)),
    ('Removed states count', len(removed_ranking)),
    ('Pomace 1L exact SKU slug', POMACE_1L),
    ('Pomace 1L calculation', 'Available Blinkit pincodes with exact SKU in_stock=1 / Blinkit serviceable pincodes in that state'),
    ('Any JIVO calculation', 'Blinkit serviceable pincodes where at least one JIVO SKU in this data layer is in stock / Blinkit serviceable pincodes'),
    ('Included states', ', '.join(states)),
    ('Removed states', ', '.join(r['state'] for r in removed_ranking)),
]
for row in source_rows:
    ws.append(row)
style_header(ws)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 120

wb.save(OUT_XLSX)
print(OUT_XLSX)
print('included', len(states), states)
print('removed', len(removed_ranking), [r['state'] for r in removed_ranking])
print('pomace top rows')
for r in filtered_ranking[:10]:
    st = r['state']
    a, t = state_sku_counts[POMACE_1L][st]
    print(r['rank'], st, r['litres'], f'{a}/{t}', f'{pct(a,t):.1%}')
