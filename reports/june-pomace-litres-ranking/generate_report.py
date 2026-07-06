import json
import pathlib
import tomllib
import urllib.parse
import urllib.request
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path('/root/pa-clients/jivo-data-bank')
OUT = ROOT / 'reports/june-pomace-litres-ranking'
OUT.mkdir(parents=True, exist_ok=True)
XLSX = OUT / 'june-pomace-litres-ranking.xlsx'
JSON_OUT = OUT / 'june-pomace-live-api-response.json'

CFG = tomllib.loads(pathlib.Path('/root/.config/jivo-ecom-pp-cli/config.toml').read_text())
BASE = CFG['base_url'].rstrip('/')
HEADERS = {'Authorization': 'Bearer ' + CFG['access_token']}
PARAMS = {
    'year': '2026',
    'month': '6',
    'metric': 'litres',
    'category': 'OLIVE',
    'sub_category': 'JIVO POMACE',
    'brand': 'JIVO',
}
URL = BASE + '/api/dashboard/state-sales?' + urllib.parse.urlencode(PARAMS)
req = urllib.request.Request(URL, headers=HEADERS)
data = json.loads(urllib.request.urlopen(req, timeout=90).read())
# Never persist or print credentials. URL contains only query params.
JSON_OUT.write_text(json.dumps(data, indent=2, ensure_ascii=False))
retrieved_at = datetime.now(timezone.utc).isoformat()

states = sorted(data.get('states') or [], key=lambda x: float(x.get('units') or x.get('value') or 0), reverse=True)
cities = sorted(data.get('cities') or [], key=lambda x: float(x.get('units') or x.get('value') or 0), reverse=True)
platforms = sorted({p for row in states + cities for p in (row.get('by_platform') or {}).keys()})

def units(row):
    return float(row.get('units') or row.get('value') or 0)

def fmt_state(s):
    return str(s).title()

def add_header(ws, headers):
    ws.append(headers)
    fill = PatternFill('solid', fgColor='111827')
    font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin', color='D1D5DB'), right=Side(style='thin', color='D1D5DB'), top=Side(style='thin', color='D1D5DB'), bottom=Side(style='thin', color='D1D5DB'))
    ws.freeze_panes = 'A2'

thin = Side(style='thin', color='D1D5DB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = 'State Ranking'
add_header(ws, ['Rank', 'State', 'Litres Sold'] + platforms)
for i, row in enumerate(states, 1):
    byp = row.get('by_platform') or {}
    ws.append([i, fmt_state(row.get('state')), units(row)] + [float(byp.get(p) or 0) for p in platforms])

ws2 = wb.create_sheet('City Ranking')
add_header(ws2, ['Rank', 'City', 'State', 'Litres Sold'] + platforms)
for i, row in enumerate(cities, 1):
    byp = row.get('by_platform') or {}
    ws2.append([i, str(row.get('city') or '').title(), fmt_state(row.get('state') or ''), units(row)] + [float(byp.get(p) or 0) for p in platforms])

ws3 = wb.create_sheet('Platform Totals')
add_header(ws3, ['Platform', 'Litres Sold'])
platform_totals = {p: 0.0 for p in platforms}
for row in states:
    for p, v in (row.get('by_platform') or {}).items():
        platform_totals[p] = platform_totals.get(p, 0.0) + float(v or 0)
for p, v in sorted(platform_totals.items(), key=lambda kv: kv[1], reverse=True):
    ws3.append([p, v])

ws4 = wb.create_sheet('Source Summary')
source_rows = [
    ('Source', 'Live Jivo ecom app/API'),
    ('Endpoint', '/api/dashboard/state-sales'),
    ('Base URL', BASE),
    ('Params', json.dumps(PARAMS, ensure_ascii=False)),
    ('Retrieved at UTC', retrieved_at),
    ('Product filter used', 'brand=JIVO, category=OLIVE, sub_category=JIVO POMACE'),
    ('Metric', data.get('metric_label') or data.get('metric') or 'litres'),
    ('Total litres', float(data.get('total_units') or 0)),
    ('Mapped litres', float(data.get('mapped_units') or 0)),
    ('Mapped %', float(data.get('pct_mapped') or 0)),
    ('State count', len(states)),
    ('City count', len(cities)),
    ('Errors', json.dumps(data.get('errors') or [], ensure_ascii=False)),
]
for row in source_rows:
    ws4.append(row)
ws4.column_dimensions['A'].width = 28
ws4.column_dimensions['B'].width = 100

for wsx in wb.worksheets:
    wsx.sheet_view.showGridLines = False
    for row in wsx.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.0'
    for c in range(1, wsx.max_column + 1):
        wsx.column_dimensions[get_column_letter(c)].width = max(12, min(28, max(len(str(wsx.cell(r, c).value or '')) for r in range(1, min(wsx.max_row, 40) + 1)) + 2))

wb.save(XLSX)
# Verify
wb2 = load_workbook(XLSX, read_only=True, data_only=True)
assert 'State Ranking' in wb2.sheetnames and 'City Ranking' in wb2.sheetnames
assert wb2['State Ranking'].max_row - 1 == len(states)
wb2.close()

summary = {
    'xlsx': str(XLSX),
    'state_count': len(states),
    'city_count': len(cities),
    'total_litres': float(data.get('total_units') or 0),
    'mapped_litres': float(data.get('mapped_units') or 0),
    'top_states': [(fmt_state(r.get('state')), units(r)) for r in states[:10]],
    'platform_totals': sorted(platform_totals.items(), key=lambda kv: kv[1], reverse=True),
}
print(json.dumps(summary, indent=2))
