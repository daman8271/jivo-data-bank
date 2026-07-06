from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SRC = Path('/root/.hermes/document_cache/doc_aa130968b235_blinkit-website-states-availability-3.xlsx')
OUT_DIR = Path('/root/pa-clients/jivo-data-bank/reports/blinkit-exact-16-states-availability')
OUT = OUT_DIR / 'blinkit-exact-16-states-availability.xlsx'
OUT_DIR.mkdir(parents=True, exist_ok=True)

STATES = [
    'Delhi',
    'Maharashtra',
    'Karnataka',
    'Haryana',
    'Uttar Pradesh',
    'Telangana',
    'West Bengal',
    'Punjab',
    'Tamil Nadu',
    'Andhra Pradesh',
    'Kerala',
    'Odisha',
    'Gujarat',
    'Rajasthan',
    'Madhya Pradesh',
    'Chandigarh',
]

wb = load_workbook(SRC)


def copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
    if src.hyperlink:
        dst._hyperlink = copy(src.hyperlink)
    if src.comment:
        dst.comment = copy(src.comment)


def replace_state_matrix(sheet_name):
    old = wb[sheet_name]
    old_index = wb.sheetnames.index(sheet_name)
    headers = [old.cell(1, c).value for c in range(1, old.max_column + 1)]
    col_by_state = {h: idx + 1 for idx, h in enumerate(headers) if h}
    missing = [s for s in STATES if s not in col_by_state]
    if missing:
        raise RuntimeError(f'{sheet_name} missing states: {missing}')

    temp_name = f'{sheet_name} Filtered'
    if temp_name in wb.sheetnames:
        del wb[temp_name]
    ws = wb.create_sheet(temp_name, old_index)

    # First column SKU + requested states in exact order.
    source_cols = [1] + [col_by_state[s] for s in STATES]
    for new_r, old_r in enumerate(range(1, old.max_row + 1), start=1):
        for new_c, old_c in enumerate(source_cols, start=1):
            copy_cell(old.cell(old_r, old_c), ws.cell(new_r, new_c))

    # Preserve dimensions and view settings.
    ws.freeze_panes = old.freeze_panes
    ws.sheet_view.showGridLines = old.sheet_view.showGridLines
    ws.row_dimensions[1].height = old.row_dimensions[1].height
    ws.column_dimensions['A'].width = old.column_dimensions['A'].width
    for c in range(2, len(STATES) + 2):
        old_c = source_cols[c - 1]
        ws.column_dimensions[get_column_letter(c)].width = old.column_dimensions[get_column_letter(old_c)].width

    del wb[sheet_name]
    ws.title = sheet_name


for sheet in ['State Percent Matrix', 'State Counts']:
    replace_state_matrix(sheet)

# Filter Long Format to exactly these states and in requested order.
if 'Long Format' in wb.sheetnames:
    old = wb['Long Format']
    old_index = wb.sheetnames.index('Long Format')
    temp_name = 'Long Format Filtered'
    if temp_name in wb.sheetnames:
        del wb[temp_name]
    ws = wb.create_sheet(temp_name, old_index)
    # Header
    for c in range(1, old.max_column + 1):
        copy_cell(old.cell(1, c), ws.cell(1, c))
    out_r = 2
    state_order = {s: i for i, s in enumerate(STATES)}
    rows = []
    for r in range(2, old.max_row + 1):
        st = old.cell(r, 2).value
        if st in state_order:
            rows.append((state_order[st], r))
    # Keep state order, SKU order inside state as in source.
    rows.sort(key=lambda x: (x[0], x[1]))
    for _, old_r in rows:
        for c in range(1, old.max_column + 1):
            copy_cell(old.cell(old_r, c), ws.cell(out_r, c))
        out_r += 1
    ws.freeze_panes = old.freeze_panes
    ws.sheet_view.showGridLines = old.sheet_view.showGridLines
    for c in range(1, old.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = old.column_dimensions[get_column_letter(c)].width
    del wb['Long Format']
    ws.title = 'Long Format'

# Update Source sheet with a note.
if 'Source' in wb.sheetnames:
    ws = wb['Source']
    r = ws.max_row + 2
    ws.cell(r, 1).value = 'Filtered for requested states'
    ws.cell(r, 2).value = ', '.join(STATES)
    ws.cell(r + 1, 1).value = 'Removed from shared workbook'
    ws.cell(r + 1, 2).value = 'Mizoram'
    for row in range(r, r + 2):
        for col in (1, 2):
            src = ws.cell(1, col)
            dst = ws.cell(row, col)
            if src.has_style:
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)

wb.save(OUT)
print(OUT)
