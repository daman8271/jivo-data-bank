import csv
import io
import json
import pathlib
import re
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path('/root/pa-clients/jivo-data-bank')
DATA_JS = ROOT / 'reports/ecom-availability-app/data.js'
OUT_DIR = ROOT / 'reports/zepto-exact-16-states-availability'
OUT_XLSX = OUT_DIR / 'zepto-exact-16-states-availability.xlsx'
PLATFORM = 'zepto'
STATES = [
    'Delhi', 'Maharashtra', 'Karnataka', 'Haryana', 'Uttar Pradesh', 'Telangana',
    'West Bengal', 'Punjab', 'Tamil Nadu', 'Andhra Pradesh', 'Kerala', 'Odisha',
    'Gujarat', 'Rajasthan', 'Madhya Pradesh', 'Chandigarh'
]
OUT_DIR.mkdir(parents=True, exist_ok=True)

text = DATA_JS.read_text()
raw = text.split('=', 1)[1].strip().rstrip(';')
data = json.loads(raw)

coverage_rows = [c for c in data['coverage'] if c.get('platform') == PLATFORM]
coverage_by_state = {c['state']: set(map(str, c.get('pincodes', []))) for c in coverage_rows}
for st in STATES:
    coverage_by_state.setdefault(st, set())

pincode_meta = data.get('pincodes', {})
pincode_city = {}
pincode_state = {}
for p, meta in pincode_meta.items():
    if isinstance(meta, dict):
        pincode_city[str(p)] = meta.get('city') or meta.get('c') or ''
        pincode_state[str(p)] = meta.get('state') or meta.get('st') or ''
for r in data['records']:
    if r.get('pl') == PLATFORM:
        p = str(r.get('p'))
        pincode_city[p] = r.get('c') or pincode_city.get(p, '')
        pincode_state[p] = r.get('st') or pincode_state.get(p, '')

skus = sorted({r.get('s') for r in data['records'] if r.get('pl') == PLATFORM and r.get('s') and r.get('s') != '__coverage__'})


def stock_value(v):
    return 1 if v in (1, True, '1', 'true', 'True') else 0


def title_from_slug(slug):
    return slug.replace('jivo-', 'JIVO ', 1).replace('-', ' ').title().replace('1L', '1L').replace('2L', '2L').replace('3L', '3L').replace('4L', '4L').replace('5L', '5L').replace('15L', '15L').replace('200Ml', '200ml')


best = {}
for r in data['records']:
    if r.get('pl') != PLATFORM or r.get('s') == '__coverage__':
        continue
    p = str(r.get('p'))
    s = r.get('s')
    stock = stock_value(r.get('stock'))
    date = str(r.get('date') or '')
    run = str(r.get('run') or '')
    rank = (stock, date, run)
    key = (p, s)
    prev = best.get(key)
    if prev is None or rank > prev['rank']:
        best[key] = {'stock': stock, 'rank': rank, 'city': r.get('c') or pincode_city.get(p, ''), 'state': r.get('st') or pincode_state.get(p, '')}

# Backfill Punjab/Ludhiana from the approved data-bank city hub when available,
# because the website data layer has no Zepto Punjab coverage row but the shared-style
# state list includes Punjab.
punjab_source = ROOT / 'ecom/locations/Ludhiana.md'
punjab_latest_date = ''
if punjab_source.exists():
    ludhiana_text = punjab_source.read_text()
    if 'platform/zepto' in ludhiana_text[:1000] or ',zepto,' in ludhiana_text:
        pin_section = ludhiana_text.split('## Pincodes', 1)[1].split('## SKUs', 1)[0]
        ludhiana_pins = set(re.findall(r'\[\[(\d{6})\]\]', pin_section))
        if ludhiana_pins:
            coverage_by_state['Punjab'] = ludhiana_pins
            for p in ludhiana_pins:
                pincode_city[p] = 'Ludhiana'
                pincode_state[p] = 'Punjab'
        if '```csv' in ludhiana_text:
            csv_block = ludhiana_text.split('```csv', 1)[1].split('```', 1)[0].strip()
            for r in csv.DictReader(io.StringIO(csv_block)):
                if r.get('platform') != PLATFORM:
                    continue
                p = str(r.get('pincode'))
                s = r.get('canonical_sku')
                if not s:
                    continue
                if s not in skus:
                    skus.append(s)
                stock = stock_value(r.get('in_stock'))
                date = str(r.get('date_ist') or '')
                run = str(r.get('run_id') or '')
                punjab_latest_date = max(punjab_latest_date, date)
                key = (p, s)
                rank = (date, run, stock)
                prev = best.get(key)
                if prev is None or rank > prev['rank']:
                    best[key] = {'stock': stock, 'rank': rank, 'city': 'Ludhiana', 'state': 'Punjab'}
skus = sorted(skus)

# City coverage from exactly the requested states.
coverage_by_city = defaultdict(set)
city_state = {}
for st in STATES:
    for p in coverage_by_state.get(st, set()):
        city = pincode_city.get(p) or 'Unknown'
        coverage_by_city[city].add(p)
        city_state[city] = st
cities = sorted(coverage_by_city.keys())

state_counts = {}
city_counts = {}
for sku in skus:
    state_counts[sku] = {}
    for st in STATES:
        pins = coverage_by_state.get(st, set())
        avail = sum(1 for p in pins if best.get((p, sku), {}).get('stock') == 1)
        state_counts[sku][st] = (avail, len(pins))
    city_counts[sku] = {}
    for city in cities:
        pins = coverage_by_city[city]
        avail = sum(1 for p in pins if best.get((p, sku), {}).get('stock') == 1)
        city_counts[sku][city] = (avail, len(pins))

wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill('solid', fgColor='111827')
header_font = Font(color='FFFFFF', bold=True)
good_fill = PatternFill('solid', fgColor='DCFCE7')
mid_fill = PatternFill('solid', fgColor='FEF3C7')
bad_fill = PatternFill('solid', fgColor='FEE2E2')
zero_fill = PatternFill('solid', fgColor='F3F4F6')
thin = Side(style='thin', color='D1D5DB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def pct(avail, total):
    return (avail / total) if total else 0


def fill_for(avail, total):
    v = pct(avail, total)
    if total == 0 or avail == 0:
        return zero_fill
    if v >= 0.8:
        return good_fill
    if v >= 0.4:
        return mid_fill
    return bad_fill


def style_matrix(ws, columns, is_pct=True):
    ws.freeze_panes = 'B2'
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        row[0].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        row[0].font = Font(bold=True)
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions['A'].width = 54
    for idx in range(2, len(columns) + 2):
        ws.column_dimensions[get_column_letter(idx)].width = 16
    ws.row_dimensions[1].height = 36
    if is_pct:
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.number_format = '0.0%'
                cell.fill = zero_fill if not cell.value else (good_fill if cell.value >= 0.8 else mid_fill if cell.value >= 0.4 else bad_fill)


def add_percent_sheet(name, columns, counts):
    ws = wb.create_sheet(name)
    ws.append(['SKU'] + columns)
    for sku in skus:
        ws.append([title_from_slug(sku)] + [pct(*counts[sku][col]) for col in columns])
    style_matrix(ws, columns, True)


def add_count_sheet(name, columns, counts):
    ws = wb.create_sheet(name)
    ws.append(['SKU'] + columns)
    for sku in skus:
        ws.append([title_from_slug(sku)] + [f'{counts[sku][col][0]}/{counts[sku][col][1]}' for col in columns])
    style_matrix(ws, columns, False)
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            a, t = map(int, str(cell.value).split('/'))
            cell.fill = fill_for(a, t)


add_percent_sheet('State Percent Matrix', STATES, state_counts)
add_count_sheet('State Counts', STATES, state_counts)
add_percent_sheet('City Percent Matrix', cities, city_counts)
add_count_sheet('City Counts', cities, city_counts)

ws = wb.create_sheet('Long Format')
ws.append(['Platform', 'State', 'City', 'SKU Slug', 'SKU', 'Available Pincodes', 'Serviceable Pincodes', 'Availability %'])
for sku in skus:
    for st in STATES:
        a, t = state_counts[sku][st]
        ws.append([PLATFORM, st, '', sku, title_from_slug(sku), a, t, pct(a, t)])
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row[7].number_format = '0.0%'
ws.freeze_panes = 'A2'
for col, width in {'A':14,'B':22,'C':18,'D':66,'E':66,'F':18,'G':20,'H':16}.items():
    ws.column_dimensions[col].width = width

ws = wb.create_sheet('Source')
source_rows = [
    ('Platform', PLATFORM),
    ('Data file used', 'reports/ecom-availability-app/data.js'),
    ('Generated from', data.get('generatedFrom')),
    ('Generated at', data.get('generatedAt')),
    ('Latest observation date', data.get('latestObservationDate')),
    ('Requested states', ', '.join(STATES)),
    ('State count', len(STATES)),
    ('SKU count', len(skus)),
    ('Calculation', 'Available pincodes / Zepto serviceable pincodes'),
    ('Serviceable denominator', 'Zepto coverage[] pincodes from reports/ecom-availability-app/data.js; Punjab backfilled from ecom/locations/Ludhiana.md if available; missing states stay 0/0'),
    ('Availability numerator', 'Latest/best Zepto rows per (pincode, SKU), in_stock=1'),
    ('Punjab source', f'ecom/locations/Ludhiana.md; latest Zepto observation date {punjab_latest_date or "not found"}; serviceable pincodes {len(coverage_by_state.get("Punjab", set()))}'),
    ('States with 0/0', ', '.join([st for st in STATES if len(coverage_by_state.get(st, set())) == 0]) or 'None'),
]
for row in source_rows:
    ws.append(row)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
for row in ws.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 120

wb.save(OUT_XLSX)
print(OUT_XLSX)
print('states', STATES)
print('zero_states', [st for st in STATES if len(coverage_by_state.get(st, set())) == 0])
print('sku_count', len(skus))
print('sheets', wb.sheetnames)
