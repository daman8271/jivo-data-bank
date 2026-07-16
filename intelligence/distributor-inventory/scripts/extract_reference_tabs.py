#!/usr/bin/env python3
"""Extract small, decision-relevant reference tabs from a Google Sheets XLSX export."""
from __future__ import annotations

import argparse
import csv
import gzip
import io
import json
from pathlib import Path
from openpyxl import load_workbook

TABS = {
    "MAPPING": "mapping.csv.gz",
    "SKU CODE LIST": "sku-code-list.csv.gz",
    "MASTER_SHEET": "master-sheet.csv.gz",
    "PO_Import_Log": "po-import-log.csv.gz",
    "CM_Import_Log": "citymall-import-log.csv.gz",
    "Sheet36": "po-grn-date-reference.csv.gz",
}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--out-dir", required=True, type=Path)
    args = parser.parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(args.xlsx, read_only=True, data_only=True)
    manifest = {"schema_version": 1, "source_workbook": str(args.xlsx), "tabs": {}}
    for tab, filename in TABS.items():
        ws = wb[tab]
        output = args.out_dir / filename
        rows = 0
        max_columns = 0
        with output.open("wb") as fh, gzip.GzipFile(filename="", fileobj=fh, mode="wb", compresslevel=9, mtime=0) as gz, io.TextIOWrapper(
            gz, encoding="utf-8", newline=""
        ) as txt:
            writer = csv.writer(txt)
            for row in ws.iter_rows(values_only=True):
                values = list(row)
                while values and values[-1] is None:
                    values.pop()
                if not values or not any(v is not None for v in values):
                    continue
                writer.writerow(["" if v is None else v for v in values])
                rows += 1
                max_columns = max(max_columns, len(values))
        manifest["tabs"][tab] = {"file": filename, "rows": rows, "max_columns": max_columns}
    (args.out_dir / "reference-tabs-manifest.json").write_text(json.dumps(manifest, indent=2) + "\n")


if __name__ == "__main__":
    main()
