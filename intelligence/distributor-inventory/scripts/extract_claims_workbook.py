#!/usr/bin/env python3
"""Extract every non-empty tab and formula from the DISTRIBUTORS CLAIMS workbook."""
from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import io
import json
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def slug(name: str) -> str:
    value = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return value or "sheet"


def open_gzip_text(path: Path):
    raw = path.open("wb")
    gz = gzip.GzipFile(filename="", fileobj=raw, mode="wb", compresslevel=9, mtime=0)
    text = io.TextIOWrapper(gz, encoding="utf-8", newline="")
    return raw, gz, text


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--out-dir", required=True, type=Path)
    parser.add_argument("--snapshot-date", required=True)
    args = parser.parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)
    values = load_workbook(args.xlsx, read_only=True, data_only=True)
    formulas = load_workbook(args.xlsx, read_only=True, data_only=False)
    manifest = {
        "schema_version": 1,
        "snapshot_date": args.snapshot_date,
        "document_title": "DISTRIBUTORS CLAIMS",
        "document_id": "1AkguFx1i3qIIwIG5GUZkdAIitu6_TgWhHfmsuKk8lWU",
        "workbook_export_sha256": hashlib.sha256(args.xlsx.read_bytes()).hexdigest(),
        "workbook_export_bytes": args.xlsx.stat().st_size,
        "sheets": {},
    }
    formula_path = args.out_dir / "formula-map.csv.gz"
    raw, gz, text = open_gzip_text(formula_path)
    formula_writer = csv.writer(text)
    formula_writer.writerow(["sheet", "cell", "formula"])
    total_formulas = 0
    try:
        for sheet_name in values.sheetnames:
            value_ws = values[sheet_name]
            formula_ws = formulas[sheet_name]
            output = args.out_dir / f"{slug(sheet_name)}.csv.gz"
            output_raw, output_gz, output_text = open_gzip_text(output)
            writer = csv.writer(output_text)
            rows = 0
            max_columns = 0
            formula_count = 0
            try:
                for row_number, (value_row, formula_row) in enumerate(
                    zip(value_ws.iter_rows(values_only=True), formula_ws.iter_rows(values_only=True)), 1
                ):
                    values_list = list(value_row)
                    while values_list and values_list[-1] is None:
                        values_list.pop()
                    if values_list and any(value is not None for value in values_list):
                        writer.writerow(["" if value is None else value for value in values_list])
                        rows += 1
                        max_columns = max(max_columns, len(values_list))
                    for column_number, formula in enumerate(formula_row, 1):
                        if isinstance(formula, str) and formula.startswith("="):
                            cell = f"{get_column_letter(column_number)}{row_number}"
                            formula_writer.writerow([sheet_name, cell, formula])
                            formula_count += 1
                            total_formulas += 1
            finally:
                output_text.close()
                output_gz.close()
                output_raw.close()
            if rows == 0:
                output.unlink()
                file_name = None
            else:
                file_name = output.name
            manifest["sheets"][sheet_name] = {
                "file": file_name,
                "nonempty_rows": rows,
                "max_columns": max_columns,
                "formula_cells": formula_count,
            }
    finally:
        text.close()
        gz.close()
        raw.close()
    manifest["formula_map"] = {"file": formula_path.name, "formula_cells": total_formulas}
    (args.out_dir / "workbook-manifest.json").write_text(json.dumps(manifest, indent=2) + "\n")


if __name__ == "__main__":
    main()
