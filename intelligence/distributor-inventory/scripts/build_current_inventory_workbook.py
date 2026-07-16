#!/usr/bin/env python3
"""Render the current expected inventory ledger as a user-facing XLSX."""
from __future__ import annotations

import argparse
import csv
import gzip
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F4E78"
BLUE = "D9EAF7"
GREEN = "E2F0D9"
AMBER = "FFF2CC"
RED = "F4CCCC"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7B7B7")


def style_sheet(ws, freeze="A4"):
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[3]:
        if cell.value is not None:
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(color=WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=THIN)
    for column in range(1, ws.max_column + 1):
        width = 12
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, column).value
            if value is not None:
                width = max(width, min(45, len(str(value)) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width


def add_title(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color=NAVY)
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color="666666")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--summary", required=True, type=Path)
    parser.add_argument("--ledger", required=True, type=Path)
    parser.add_argument("--baseline-coverage", required=True, type=Path)
    parser.add_argument("--source-registry", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    args = parser.parse_args()
    summary = json.loads(args.summary.read_text())
    coverage = json.loads(args.baseline_coverage.read_text())
    sources = json.loads(args.source_registry.read_text())
    with gzip.open(args.ledger, "rt", encoding="utf-8", newline="") as handle:
        ledger_rows = list(csv.DictReader(handle))

    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "Current Summary"
    add_title(ws, "Distributor Inventory — Latest Available", "Expected inventory as of close 16 July 2026; confidence and source limitations are explicit.")
    headers = ["Distributor", "Baseline cutoff", "Confidence", "Opening", "Billing inward", "Confirmed outward", "Expected", "Negative SKUs", "Interpretation"]
    ws.append(headers)
    for distributor, values in summary["current_expected_inventory"].items():
        confidence = values["baseline_confidence"]
        interpretation = "Expected from dated physical baseline" if confidence.startswith("physical") else "Provisional expected from manual opening"
        ws.append([
            distributor, values["baseline_cutoff"], confidence, values["opening_quantity"], values["billing_inward_proxy"],
            values["confirmed_outward"], values["expected_quantity"], values["negative_expected_skus"], interpretation,
        ])
    for distributor, values in summary["movement_only"].items():
        ws.append([distributor, None, "movement_only", None, values["billing_inward_proxy"], values["confirmed_outward"], None, None, values["reason"]])
    for row in ws.iter_rows(min_row=4):
        confidence = row[2].value
        fill = GREEN if confidence and str(confidence).startswith("physical") else AMBER
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
    style_sheet(ws)

    ws = wb.create_sheet("SKU Ledger")
    add_title(ws, "SKU-level expected inventory", "Accounting-negative rows are diagnostic and excluded from physical movement.")
    headers = list(ledger_rows[0]) if ledger_rows else []
    ws.append(headers)
    numeric_headers = {
        "opening_quantity", "billing_inward_proxy", "accounting_negative_diagnostic",
        "confirmed_outward", "expected_quantity",
    }
    for row in ledger_rows:
        values = []
        for header in headers:
            value = row[header]
            if header in numeric_headers and value != "":
                value = float(value)
            values.append(value)
        ws.append(values)
    for row in ws.iter_rows(min_row=4):
        if row[headers.index("status")].value == "expected_negative":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=RED)
    style_sheet(ws)

    ws = wb.create_sheet("Baseline Coverage")
    add_title(ws, "Current baseline coverage", "Only dated or explicitly assumed openings can seed absolute inventory.")
    ws.append(["Distributor", "SKU rows", "Baseline total", "Cutoff", "Confidence"])
    for distributor, values in coverage["by_distributor"].items():
        ws.append([distributor, values["canonical_sku_rows"], values["baseline_total_units"], values["effective_cutoff"], values["confidence"]])
    ws.append([])
    ws.append(["Not currently seeded"])
    for distributor in coverage["not_currently_seeded"]:
        ws.append([distributor])
    ws.append([])
    ws.append(["Notes"])
    for note in coverage["notes"]:
        ws.append([note])
    style_sheet(ws)

    ws = wb.create_sheet("Source Status")
    add_title(ws, "Source status", "All three shared Google Sheets are now publicly readable and registered.")
    ws.append(["Source", "Document", "Purpose", "Access", "Rows/date coverage", "Inventory role"])
    for source in sources["sources"]:
        coverage_text = source.get("rows") or source.get("important_date_ranges") or source.get("target_tab_rows")
        ws.append([
            source["source_id"], source.get("document_title"), source.get("purpose"), source.get("access"),
            json.dumps(coverage_text) if isinstance(coverage_text, dict) else coverage_text,
            source.get("inventory_role") or source.get("latest_precedence") or "See source rules",
        ])
    style_sheet(ws)

    ws = wb.create_sheet("Rules & Limitations")
    add_title(ws, "Rules and limitations", "No balancing plugs or unsupported physical movements are used.")
    ws.append(["Type", "Rule"])
    ws.append(["Equation", summary["equation"]])
    for limitation in summary["limitations"]:
        ws.append(["Limitation", limitation])
    ws.append(["Rule", "Use confirmed delivered quantity by delivery date; ordered or unfulfilled quantity is not movement."])
    ws.append(["Rule", "Platform/customer GRN is outward evidence but does not prove physical dispatch timing or in-transit stock."])
    ws.append(["Rule", "Live authenticated JIVO Ecom data takes precedence for the latest state; Google sheets remain auditable references."])
    style_sheet(ws)

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)


if __name__ == "__main__":
    main()
