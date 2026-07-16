#!/usr/bin/env python3
"""Build the current canonical-SKU baseline registry from preserved reconciliation sources."""
from __future__ import annotations

import argparse
import csv
import gzip
import io
import json
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

FIELDS = [
    "distributor", "sap_sku_code", "product", "baseline_quantity", "effective_cutoff",
    "baseline_type", "confidence", "source_file", "source_sheet", "source_row", "notes",
]


def writer_for(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    raw = path.open("wb")
    gz = gzip.GzipFile(filename="", fileobj=raw, mode="wb", compresslevel=9, mtime=0)
    text = io.TextIOWrapper(gz, encoding="utf-8", newline="")
    writer = csv.DictWriter(text, fieldnames=FIELDS)
    writer.writeheader()
    return raw, gz, text, writer


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo-root", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    parser.add_argument("--coverage", required=True, type=Path)
    args = parser.parse_args()
    root = args.repo_root
    rows: list[dict] = []

    antize_file = root / "intelligence/distributor-inventory/derived/reconciliations/antize-02jul-to-15jul-2026.xlsx"
    ws = load_workbook(antize_file, read_only=True, data_only=True)["Reconciliation"]
    for row_number, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
        code = row[0]
        if not isinstance(code, str) or not code.startswith("FG"):
            continue
        rows.append({
            "distributor": "ANTIZE FOODS PRIVATE LIMITED",
            "sap_sku_code": code,
            "product": row[2],
            "baseline_quantity": row[10],
            "effective_cutoff": "2026-07-15",
            "baseline_type": "physical_closing",
            "confidence": "physical_verified",
            "source_file": str(antize_file.relative_to(root)),
            "source_sheet": "Reconciliation",
            "source_row": row_number,
            "notes": "Canonical mapped physical closing; three blank source quantities remain controlled unknowns outside populated total.",
        })

    baba_file = root / "intelligence/distributor-inventory/derived/reconciliations/baba-01jul-to-16jul-2026.xlsx"
    ws = load_workbook(baba_file, read_only=True, data_only=True)["Reconciliation"]
    for row_number, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        code = row[0]
        if not isinstance(code, str) or not code.startswith("FG"):
            continue
        rows.append({
            "distributor": "BABA LOKENATH TRADERS",
            "sap_sku_code": code,
            "product": row[1],
            "baseline_quantity": row[6],
            "effective_cutoff": "2026-07-16",
            "baseline_type": "physical_closing",
            "confidence": "physical_verified_canonical_scope",
            "source_file": str(baba_file.relative_to(root)),
            "source_sheet": "Reconciliation",
            "source_row": row_number,
            "notes": "Canonical scope. Four stable alternate-pack/carton rows totaling 166 units are tracked outside this canonical ledger.",
        })

    central_file = root / "intelligence/distributor-inventory/raw/2026-07-16/physical-sources/manual-dis-stock-report.xlsx"
    ws = load_workbook(central_file, read_only=True, data_only=True)["Sheet1"]
    distributor_columns = {
        "CHIRAG ENTERPRISES MUMBAI": (5, "physical_statement_mapped", "physical_statement_control"),
        "SUSTAINQUEST PRIVATE LIMITED": (21, "manual_tracker_unverified", "manual_tracker_opening"),
    }
    for row_number, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
        code = row[0]
        if not isinstance(code, str) or not code.startswith("FG"):
            continue
        for distributor, (column_index, confidence, baseline_type) in distributor_columns.items():
            quantity = row[column_index] if len(row) > column_index else None
            if quantity is None:
                continue
            rows.append({
                "distributor": distributor,
                "sap_sku_code": code,
                "product": row[1],
                "baseline_quantity": quantity,
                "effective_cutoff": "2026-06-30",
                "baseline_type": baseline_type,
                "confidence": confidence,
                "source_file": str(central_file.relative_to(root)),
                "source_sheet": "Sheet1",
                "source_row": row_number,
                "notes": "Chirag cutoff is supported by the June statement; SustainQuest cutoff is the July monthly-opening operational assumption and remains provisional.",
            })

    rows.sort(key=lambda item: (item["distributor"], item["sap_sku_code"]))
    raw, gz, text, writer = writer_for(args.out)
    try:
        writer.writerows(rows)
    finally:
        text.close()

    totals = defaultdict(float)
    counts = defaultdict(int)
    confidence = {}
    cutoff = {}
    for row in rows:
        totals[row["distributor"]] += float(row["baseline_quantity"] or 0)
        counts[row["distributor"]] += 1
        confidence[row["distributor"]] = row["confidence"]
        cutoff[row["distributor"]] = row["effective_cutoff"]
    coverage = {
        "schema_version": 1,
        "as_of_date": "2026-07-16",
        "baseline_rows": len(rows),
        "by_distributor": {
            distributor: {
                "canonical_sku_rows": counts[distributor],
                "baseline_total_units": totals[distributor],
                "effective_cutoff": cutoff[distributor],
                "confidence": confidence[distributor],
            }
            for distributor in sorted(totals)
        },
        "not_currently_seeded": [
            "KNOWTABLE ONLINE SERVICES PRIVATE LIMITED",
            "EVARA ENTERPRISES"
        ],
        "notes": [
            "Baba canonical total excludes four stable alternate-pack/carton rows totaling 166 units.",
            "SustainQuest uses the manual July opening assumption until a dated physical snapshot is available.",
            "Knowtable and Evara have movements but no dated July opening baseline.",
            "Undated DISTRIBUTORS CLAIMS stock tabs are historical and are not used as current baselines."
        ]
    }
    args.coverage.parent.mkdir(parents=True, exist_ok=True)
    args.coverage.write_text(json.dumps(coverage, indent=2) + "\n")


if __name__ == "__main__":
    main()
